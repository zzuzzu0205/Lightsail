from datetime import datetime
import os
from enum import auto

import pandas as pd
import openpyxl
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from django.http import HttpResponseRedirect

from django.shortcuts import render

# Create your views here.
from django.urls import reverse
from openpyxl.styles import PatternFill, Alignment

from mainapp.models import Category, Review, SecondLabeledData


def output(request):
    try:

        # category_product 변수를 get 방식 으로 받으면 세션에 저장
        if request.method == "GET":
            if 'category_product' in request.GET:
                request.session['category_product'] = request.GET['category_product']
                print(request.session['category_product'])
                request.session.set_expiry(300)

                category_product = request.session['category_product']
                category_detail = Category.objects.filter(category_product=category_product)
                alltotal = Review.objects.filter(category_product=category_product).count()
                first_num = Review.objects.filter(category_product=category_product).filter(first_status=True).count()
                second_num = Review.objects.filter(category_product=category_product).filter(second_status=True).count()
                dummy_num = Review.objects.filter(category_product=category_product).filter(dummy_status=True).count()

                context = {'category_detail': category_detail}
                context['alltotal'] = alltotal
                context['first_num'] = first_num
                context['dummy_num'] = dummy_num
                context['second_num'] = second_num
                return render(request, 'outputapp/output.html', context)

            else:
                context = {'message': '제품을 다시 선택해주세요.'}
                return render(request, 'outputapp/output.html', context)

        elif request.method == "POST" and 'export' in request.POST:
            if 'export' in '.xlsx export':
                print("엑셀")
                product = request.POST['product']
                print(product)
                category = Category.objects.filter(category_product=product)
                category_list = []
                category_id_list = []
                for i in category:
                    category_list.append(i.category_middle)
                    category_id_list.append(i.category_id)

                print("카테고리리스트", category_list)
                print("카테고리id리스트", category_id_list)
                category_len = len(category_list)

                # 워크북(엑셀파일)을 새로 만듭니다.
                wb = openpyxl.Workbook()
                wb.active.title = category_list[0]
                sheet = wb.active
                sheet.append(["Product_Group", "Category", "긍정 키워드", "부정 키워드", "중립 키워드"])

                # 음영 색 지정
                grayFill = PatternFill(start_color='BFBFBF',
                                         end_color='BFBFBF',
                                         fill_type='solid')
                # 지정된 음영 색으로 영역 색칠하기
                sheet["A1"].fill = grayFill
                sheet["B1"].fill = grayFill
                sheet["C1"].fill = grayFill
                sheet["D1"].fill = grayFill
                sheet["E1"].fill = grayFill

                # 중앙 정렬 및 너비 설정
                format1 = Alignment(horizontal='center', vertical='center')
                sheet["A1"].alignment = format1
                sheet["B1"].alignment = format1
                sheet["C1"].alignment = format1
                sheet["D1"].alignment = format1
                sheet["E1"].alignment = format1
                sheet.column_dimensions['A'].width = 13
                sheet.column_dimensions['B'].width = 13
                sheet.column_dimensions['C'].width = 25
                sheet.column_dimensions['D'].width = 25
                sheet.column_dimensions['E'].width = 25

                data_couple = SecondLabeledData.objects.filter(category_id=category_id_list[0])
                s = 2
                f = 2
                i = 2
                j = 2
                k = 2

                for data in data_couple: # 긍정, 부정, 중립 키워드 수 중 가장 큰것에 맞춰서 제품 수와 카테고리 뽑아내야함
                    print(product)
                    p = product
                    sheet.cell(row=s, column=1).value = p
                    s = s + 1
                    print(category_list[0])
                    c = category_list[0]
                    sheet.cell(row=f, column=2).value = c
                    f = f + 1

                for data in data_couple:
                    print(i, data)
                    if data.second_labeled_emotion == 'positive':
                        t = data.second_labeled_target + ' AND ' + data.second_labeled_expression
                        sheet.cell(row=i, column=3).value = t
                        i = i + 1
                    elif data.second_labeled_emotion == 'negative':
                        t = data.second_labeled_target + ' AND ' + data.second_labeled_expression
                        sheet.cell(row=j, column=4).value = t
                        j = j + 1
                    elif data.second_labeled_emotion == 'neutral':
                        t = data.second_labeled_target + ' AND ' + data.second_labeled_expression
                        sheet.cell(row=k, column=5).value = t
                        k = k + 1
                    else:
                        print('해당없음')

                for ii in range(1, category_len):
                    category_list[ii] = wb.create_sheet("%s" % category_list[ii])
                    sheet = category_list[ii]
                    # 헤더 추가하기
                    sheet.append(["Product_Group", "Category", "긍정 키워드", "부정 키워드", "중립 키워드"])

                    # 음영 색 지정
                    grayFill = PatternFill(start_color='BFBFBF',
                                           end_color='BFBFBF',
                                           fill_type='solid')
                    # 지정된 음영 색으로 음역 색칠하기
                    sheet["A1"].fill = grayFill
                    sheet["B1"].fill = grayFill
                    sheet["C1"].fill = grayFill
                    sheet["D1"].fill = grayFill
                    sheet["E1"].fill = grayFill

                    # 중앙 정렬 및 너비 설정
                    format1 = Alignment(horizontal='center', vertical='center')
                    sheet["A1"].alignment = format1
                    sheet["B1"].alignment = format1
                    sheet["C1"].alignment = format1
                    sheet["D1"].alignment = format1
                    sheet["E1"].alignment = format1
                    sheet.column_dimensions['A'].width = 13
                    sheet.column_dimensions['B'].width = 13
                    sheet.column_dimensions['C'].width = 25
                    sheet.column_dimensions['D'].width = 25
                    sheet.column_dimensions['E'].width = 25

                    data_couple = SecondLabeledData.objects.filter(category_id=category_id_list[ii])
                    s = 2
                    f = 2
                    i = 2
                    j = 2
                    k = 2

                    for data in data_couple:  # 긍정, 부정, 중립 키워드 수 중 가장 큰것에 맞춰서 제품 수와 카테고리 뽑아내야함
                        print(product)
                        p = product
                        sheet.cell(row=s, column=1).value = p
                        s = s + 1
                        print(category_list[0])
                        c = category_list[0]
                        sheet.cell(row=f, column=2).value = c
                        f = f + 1

                    for data in data_couple:
                        print(i, data)
                        if data.second_labeled_emotion == 'positive':
                            t = data.second_labeled_target + ' AND ' + data.second_labeled_expression
                            sheet.cell(row=i, column=3).value = t
                            i = i + 1
                        elif data.second_labeled_emotion == 'negative':
                            t = data.second_labeled_target + ' AND ' + data.second_labeled_expression
                            sheet.cell(row=j, column=4).value = t
                            j = j + 1
                        elif data.second_labeled_emotion == 'neutral':
                            t = data.second_labeled_target + ' AND ' + data.second_labeled_expression
                            sheet.cell(row=k, column=5).value = t
                            k = k + 1
                        else:
                            print('해당없음')


                # 제품명 + 날짜 + 시간으로 파일명이 생성되며 c드라이브에 labelingresult 폴더 없을시 자동 생성된 후 그 안에 파일이 담김

                Path("C:\\labelingresult").mkdir(exist_ok=True)
                filename = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
                filedir = "C:\\labelingresult\\"
                wb.save(filedir + product + '_' + filename + '.xlsx')


            elif 'export' in '.data analysis':
                print("분석")

            else:
                print("에러")

        else:
            print("에러")


    except Exception as identifier:
        print(identifier)
    return render(request, 'outputapp/output.html')
